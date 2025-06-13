# import module
import openpyxl
import os
from pathlib import Path
from processFolder import process_files_in_folder
import json
from io import StringIO
import csv

def processUploadedFile(uploadedFile):
# load excel with its path
    print("Uploaded file ", uploadedFile)
    wrkbk = openpyxl.load_workbook(uploadedFile)

    sh_tripList = wrkbk['Trip_List']
    sh_evidence = wrkbk['Evidence']
    sh_members = wrkbk['Trip_Members']
    completeExtractedData = ""
    # iterate through excel and display data
    for i in range(2, sh_tripList.max_row+1):
        print("\n")
        print("Row ", i, " data :")
        
        # Get the Trip Id
        tripId = sh_tripList.cell(row=i, column=2).value
        print("Trip Id: ", tripId)
        vacationDestination = sh_tripList.cell(row=i, column=3).value
        print("Vacation Destination : ", vacationDestination)
        #Get the From Date and To Date
        fromDate = sh_tripList.cell(row=i, column=5).value
        print("From Date: ", fromDate)
        toDate = sh_tripList.cell(row=i, column=6).value
        print("To Date: ", toDate)
        for evidenceData in range(2, sh_evidence.max_row+1):  
            # Get the Trip Id
            tripIdEvidenceSheet = sh_evidence.cell(row=evidenceData, column=2).value
            if tripId == tripIdEvidenceSheet :
                #print("Trip Id: ", tripId ," with ", tripIdEvidenceSheet)
                #Get the Path of evidence
                evidencePath = str(sh_evidence.cell(row=evidenceData, column=3).value)
                folder_path = os.path.dirname(evidencePath)
                print("Evidence Path: ", folder_path)
                #process_files_in_folder(folder_path)
                break
            else :
                continue
        memberCountData = 0
        for memberData in range(2, sh_members.max_row+1):  
            # Get the Trip Id
            tripIdMemberSheet = sh_members.cell(row=memberData, column=2).value
            if tripId == tripIdMemberSheet :
                #print("Trip Id: ", tripId ," with ", tripIdEvidenceSheet)
                #Get the Path of evidence
                memberCountData = memberCountData + 1
            else :
                continue
        print("Total Member Count: " , memberCountData)
        if memberCountData > 0:
            completeExtractedData = str(i) + "," + str(tripId) + "," + str(fromDate) + "," + str(toDate) + "," + str(folder_path) + "," + str(memberCountData)
            #write to JSON file
            jsonFilePath = "jsonFiles/" + str(tripId) + ".json"
            fieldnames = ['rowId','tripId', 'fromDate', 'toDate','evidenceFolderPath','memberCount']
            print("The completeExtractedData is :" , completeExtractedData)    
            csv_string_noheaders_to_json(completeExtractedData, jsonFilePath,fieldnames)

def csv_string_noheaders_to_json(csv_string, json_file_path, fieldnames, delimiter=','):
    """
    Convert CSV string without headers to JSON file
    
    Args:
        csv_string (str): CSV data without headers
        json_file_path (str): Output file path
        fieldnames (list): List of column names
        delimiter (str): Value separator (default ',')
    """
    csv_data = StringIO(csv_string)
    reader = csv.reader(csv_data, delimiter=delimiter)
    
    data = []
    for row in reader:
        if len(row) == len(fieldnames):  # Only process valid rows
            data.append(dict(zip(fieldnames, row)))
    
    with open(json_file_path, 'w') as json_file:
        json.dump(data, json_file, indent=4)
    
    print(f"Converted CSV to {json_file_path}")


